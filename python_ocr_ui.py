"""
GLM-4V-Flash OCR 识别工具 - GUI 版本
使用智谱AI GLM-4V模型识别图片中的医学参数
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import os
import time
import threading
from queue import Queue
import base64
import io
import json
import re
from pathlib import Path
from zai import ZhipuAiClient
import pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font

# 需要提取的参数列表
PARAMETERS = [
    "IVS;d", "LVID;d", "LVPW;d",
    "IVS;s", "LVID;s", "LVPW;s",
    "LV Vol;d", "LV Vol;s", "EF", "FS",
    "LV Mass", "LV Mass Cor"
]


class OCRProcessorUI:
    # 配色方案
    COLORS = {
        'bg_main': '#F5F7FA',
        'bg_panel': '#FFFFFF',
        'primary': '#4A90D9',
        'primary_hover': '#357ABD',
        'success': '#5CB85C',
        'success_hover': '#4AA84A',
        'danger': '#E74C3C',
        'danger_hover': '#D0342C',
        'warning': '#F39C12',
        'warning_hover': '#E8910C',
        'info': '#5BC0DE',
        'text_primary': '#2C3E50',
        'text_secondary': '#7F8C8D',
        'border': '#E1E8ED',
        'border_light': '#ECF0F1',
        'input_bg': '#FAFBFC',
    }

    def __init__(self, root):
        self.root = root
        self.root.title("OCR 识别工具 - GLM-4V-Flash")
        self.root.geometry("950x750")
        self.root.minsize(800, 600)
        self.root.configure(bg=self.COLORS['bg_main'])

        # API Key
        self.api_key = tk.StringVar(value=os.environ.get("ZHIPU_API_KEY", ""))
        # 输入文件夹路径
        self.input_folder = tk.StringVar()
        # 输出文件路径
        self.output_file = tk.StringVar(value="D:/code/image/result/ocr_result.xlsx")

        # 文件夹路径列表
        self.folder_paths_list = []

        # 线程相关
        self.processing = False
        self.process_thread = None
        self.log_queue = Queue()

        self.create_widgets()
        self.check_log_queue()

    def create_widgets(self):
        # 主容器
        main_container = tk.Frame(self.root, bg=self.COLORS['bg_main'])
        main_container.pack(fill="both", expand=True, padx=16, pady=16)

        # 标题区域
        title_frame = tk.Frame(main_container, bg=self.COLORS['bg_main'])
        title_frame.grid(row=0, column=0, columnspan=2, pady=(0, 12), sticky="ew")

        title_container = tk.Frame(title_frame, bg=self.COLORS['primary'], height=50)
        title_container.pack(fill="x")
        title_container.pack_propagate(False)

        title_label = tk.Label(
            title_container,
            text="  🔍 OCR 识别工具 - GLM-4V-Flash",
            font=("Microsoft YaHei", 16, "bold"),
            bg=self.COLORS['primary'],
            fg="white",
            anchor="w"
        )
        title_label.pack(fill="both", expand=True, padx=16, pady=8)

        # 左侧面板 - 配置区域
        left_panel = tk.Frame(main_container, bg=self.COLORS['bg_panel'], relief="flat", bd=0)
        left_panel.grid(row=1, column=0, sticky="nsew", padx=(0, 10))

        # 右侧面板 - 日志
        right_panel = tk.Frame(main_container, bg=self.COLORS['bg_main'])
        right_panel.grid(row=1, column=1, sticky="nsew")

        # 配置 grid 权重
        main_container.columnconfigure(1, weight=1)
        main_container.rowconfigure(1, weight=1)

        # 左侧面板内容容器
        left_content = tk.Frame(left_panel, bg=self.COLORS['bg_panel'])
        left_content.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        left_panel.columnconfigure(0, weight=1)
        left_panel.rowconfigure(0, weight=1)

        # ============ 左侧面板 ============
        row_idx = 0

        # API Key 配置区域
        api_frame = tk.LabelFrame(
            left_content,
            text=" 🔑 API 配置 ",
            font=("Microsoft YaHei", 10, "bold"),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_primary'],
            relief="solid",
            bd=1,
            highlightbackground=self.COLORS['border'],
            highlightthickness=1
        )
        api_frame.grid(row=row_idx, column=0, sticky="ew", pady=(0, 12))
        row_idx += 1

        api_label = tk.Label(api_frame, text="智谱 AI API Key:", font=("Microsoft YaHei", 10),
                            bg=self.COLORS['bg_panel'], fg=self.COLORS['text_secondary'])
        api_label.pack(anchor="w", padx=12, pady=(10, 6))

        api_entry_frame = tk.Frame(api_frame, bg=self.COLORS['bg_panel'])
        api_entry_frame.pack(fill="x", padx=12, pady=(0, 10))

        api_entry = tk.Entry(
            api_entry_frame,
            textvariable=self.api_key,
            font=("Microsoft YaHei", 10),
            bg=self.COLORS['input_bg'],
            fg=self.COLORS['text_primary'],
            insertbackground=self.COLORS['primary'],
            relief="solid",
            bd=1,
            highlightthickness=1,
            highlightbackground=self.COLORS['border'],
            highlightcolor=self.COLORS['primary'],
            show="*"
        )
        api_entry.pack(side="left", fill="x", expand=True)

        # 输入文件夹配置
        input_frame = tk.LabelFrame(
            left_content,
            text=" 📁 输入设置 ",
            font=("Microsoft YaHei", 10, "bold"),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_primary'],
            relief="solid",
            bd=1,
            highlightbackground=self.COLORS['border'],
            highlightthickness=1
        )
        input_frame.grid(row=row_idx, column=0, sticky="ew", pady=(0, 12))
        row_idx += 1

        input_label = tk.Label(input_frame, text="选择包含图片的文件夹:",
                             font=("Microsoft YaHei", 10),
                             bg=self.COLORS['bg_panel'], fg=self.COLORS['text_secondary'])
        input_label.pack(anchor="w", padx=12, pady=(10, 6))

        # 文件夹列表框容器
        listbox_container = tk.Frame(input_frame, bg=self.COLORS['bg_panel'])
        listbox_container.pack(fill="x", padx=12, pady=(0, 8))

        scrollbar = tk.Scrollbar(
            listbox_container,
            bg=self.COLORS['bg_panel'],
            troughcolor=self.COLORS['border_light'],
            activebackground=self.COLORS['primary']
        )
        scrollbar.pack(side="right", fill="y")

        self.folder_listbox = tk.Listbox(
            listbox_container,
            font=("Microsoft YaHei", 9),
            yscrollcommand=scrollbar.set,
            selectmode=tk.MULTIPLE,
            bg=self.COLORS['input_bg'],
            fg=self.COLORS['text_primary'],
            selectbackground=self.COLORS['primary'],
            selectforeground="white",
            relief="solid",
            bd=1,
            height=4
        )
        self.folder_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.folder_listbox.yview)

        # 按钮区域
        btn_frame_inner = tk.Frame(input_frame, bg=self.COLORS['bg_panel'])
        btn_frame_inner.pack(fill="x", padx=12, pady=(0, 10))

        btn_style_common = {
            "font": ("Microsoft YaHei", 9),
            "cursor": "hand2",
            "relief": "flat",
            "bd": 0,
            "height": 1
        }

        btn_add = tk.Button(
            btn_frame_inner,
            text="➕ 添加",
            command=self.add_folder,
            bg=self.COLORS['primary'],
            fg="white",
            activebackground=self.COLORS['primary_hover'],
            **btn_style_common
        )
        btn_add.pack(side="left", fill="x", expand=True, padx=(0, 4))

        btn_remove = tk.Button(
            btn_frame_inner,
            text="➖ 移除",
            command=self.remove_folder,
            bg=self.COLORS['danger'],
            fg="white",
            activebackground=self.COLORS['danger_hover'],
            **btn_style_common
        )
        btn_remove.pack(side="left", fill="x", expand=True, padx=4)

        btn_clear = tk.Button(
            btn_frame_inner,
            text="🗑️ 清空",
            command=self.clear_folders,
            bg=self.COLORS['warning'],
            fg="white",
            activebackground=self.COLORS['warning_hover'],
            **btn_style_common
        )
        btn_clear.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # 输出文件配置
        output_frame = tk.LabelFrame(
            left_content,
            text=" 📄 输出设置 ",
            font=("Microsoft YaHei", 10, "bold"),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_primary'],
            relief="solid",
            bd=1,
            highlightbackground=self.COLORS['border'],
            highlightthickness=1
        )
        output_frame.grid(row=row_idx, column=0, sticky="ew", pady=(0, 12))
        row_idx += 1

        output_label = tk.Label(output_frame, text="输出 Excel 文件:", font=("Microsoft YaHei", 10),
                              bg=self.COLORS['bg_panel'], fg=self.COLORS['text_secondary'])
        output_label.pack(anchor="w", padx=12, pady=(10, 6))

        output_entry_frame = tk.Frame(output_frame, bg=self.COLORS['bg_panel'])
        output_entry_frame.pack(fill="x", padx=12, pady=(0, 10))

        output_entry = tk.Entry(
            output_entry_frame,
            textvariable=self.output_file,
            font=("Microsoft YaHei", 10),
            bg=self.COLORS['input_bg'],
            fg=self.COLORS['text_primary'],
            insertbackground=self.COLORS['primary'],
            relief="solid",
            bd=1,
            highlightthickness=1,
            highlightbackground=self.COLORS['border'],
            highlightcolor=self.COLORS['primary']
        )
        output_entry.pack(side="left", fill="x", expand=True)

        btn_browse = tk.Button(
            output_entry_frame,
            text="浏览",
            command=self.browse_output_file,
            width=6,
            font=("Microsoft YaHei", 9),
            bg=self.COLORS['info'],
            fg="white",
            cursor="hand2",
            relief="flat",
            activebackground=self.COLORS['primary'],
            activeforeground="white"
        )
        btn_browse.pack(side="left", padx=(6, 0))

        # 操作按钮区域
        action_frame = tk.Frame(left_content, bg=self.COLORS['bg_panel'])
        action_frame.grid(row=row_idx, column=0, sticky="ew", pady=(0, 12))
        row_idx += 1

        self.btn_process = tk.Button(
            action_frame,
            text="▶ 开始识别",
            command=self.process_images,
            bg=self.COLORS['success'],
            fg="white",
            font=("Microsoft YaHei", 12, "bold"),
            height=2,
            cursor="hand2",
            relief="flat",
            activebackground=self.COLORS['success_hover'],
            activeforeground="white",
            bd=0
        )
        self.btn_process.pack(fill="x", pady=(4, 4))

        # 说明文字
        info_text = "📋 提取参数: " + ", ".join(PARAMETERS[:6]) + "..."
        info_label = tk.Label(
            left_content,
            text=info_text,
            font=("Microsoft YaHei", 8),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_secondary'],
            wraplength=250
        )
        info_label.grid(row=row_idx, column=0, sticky="w", pady=(0, 8))

        # ============ 右侧面板 ============
        # 输出日志区域
        log_frame = tk.LabelFrame(
            right_panel,
            text=" 📋 处理日志 ",
            font=("Microsoft YaHei", 10, "bold"),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_primary'],
            relief="solid",
            bd=1,
            highlightbackground=self.COLORS['border'],
            highlightthickness=1
        )
        log_frame.grid(row=0, column=0, sticky="nsew")

        right_panel.rowconfigure(0, weight=1)
        right_panel.columnconfigure(0, weight=1)

        # 配置日志文本标签颜色
        self.text_output = scrolledtext.ScrolledText(
            log_frame,
            font=("Consolas", 10),
            wrap="word",
            bg="#1E1E1E",
            fg="#D4D4D4",
            insertbackground="#FFFFFF",
            relief="solid",
            bd=1,
            highlightthickness=0
        )

        # 配置日志颜色标签
        self.text_output.tag_config("info", foreground="#5BC0DE")
        self.text_output.tag_config("success", foreground="#5CB85C")
        self.text_output.tag_config("error", foreground="#E74C3C")
        self.text_output.tag_config("warning", foreground="#F39C12")

        self.text_output.pack(fill="both", expand=True, padx=10, pady=(8, 10))

    # ============ 线程相关方法 ============

    def log_message(self, message, tag=None):
        """线程安全的日志输出方法"""
        self.log_queue.put((message, tag))

    def check_log_queue(self):
        """检查日志队列并更新UI（在主线程中调用）"""
        try:
            while True:
                message, tag = self.log_queue.get_nowait()
                if tag:
                    self.text_output.insert(tk.END, message, tag)
                else:
                    self.text_output.insert(tk.END, message)
                self.text_output.see(tk.END)
        except:
            pass
        self.root.after(100, self.check_log_queue)

    def set_processing_state(self, processing):
        """设置处理状态"""
        self.processing = processing
        if processing:
            self.btn_process.config(text="⏳ 识别中...", state="disabled", bg=self.COLORS['text_secondary'])
        else:
            self.btn_process.config(text="▶ 开始识别", state="normal", bg=self.COLORS['success'])

    # ============ 文件夹操作方法 ============

    def add_folder(self):
        root = tk.Tk()
        root.withdraw()
        folder = filedialog.askdirectory(title="选择包含图片的文件夹")
        root.destroy()

        if folder:
            # 检查是否包含子文件夹
            subfolders = []
            for item in os.listdir(folder):
                item_path = os.path.join(folder, item)
                if os.path.isdir(item_path):
                    subfolders.append(item_path)

            if subfolders:
                msg = f"发现 {len(subfolders)} 个子文件夹，是否全部添加？\n\n"
                msg += "是 - 添加所有子文件夹\n"
                msg += "否 - 只添加选中的文件夹"
                choice = messagebox.askyesno("确认", msg)

                if choice:
                    for subfolder in subfolders:
                        if subfolder not in self.folder_paths_list:
                            self.folder_paths_list.append(subfolder)
                            self.folder_listbox.insert(tk.END, subfolder)
                else:
                    if folder not in self.folder_paths_list:
                        self.folder_paths_list.append(folder)
                        self.folder_listbox.insert(tk.END, folder)
            else:
                if folder not in self.folder_paths_list:
                    self.folder_paths_list.append(folder)
                    self.folder_listbox.insert(tk.END, folder)

    def remove_folder(self):
        selected = self.folder_listbox.curselection()
        if selected:
            for index in reversed(selected):
                self.folder_listbox.delete(index)
                del self.folder_paths_list[index]

    def clear_folders(self):
        self.folder_listbox.delete(0, tk.END)
        self.folder_paths_list.clear()

    def browse_output_file(self):
        file_path = filedialog.asksaveasfilename(
            title="保存输出文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.output_file.set(file_path)

    # ============ OCR 处理方法 ============

    def encode_image_to_base64(self, image_path):
        """将图片文件编码为base64格式"""
        try:
            with open(image_path, "rb") as image_file:
                if image_path.lower().endswith(('.tif', '.tiff')):
                    with Image.open(image_path) as img:
                        if img.mode in ('RGBA', 'P'):
                            img = img.convert('RGB')
                        buffered = io.BytesIO()
                        img.save(buffered, format="JPEG", quality=85)
                        image_data = buffered.getvalue()
                        return base64.b64encode(image_data).decode('utf-8'), "image/jpeg"
                else:
                    return base64.b64encode(image_file.read()).decode('utf-8'), "image/jpeg"
        except Exception as e:
            self.log_message(f"编码图片失败 {image_path}: {e}\n", "error")
            return None, None

    def call_glm4v_api(self, base64_image, mime_type="image/jpeg"):
        """调用智谱GLM-4V-Flash API进行图片OCR识别"""
        try:
            client = ZhipuAiClient(
                api_key=self.api_key.get(),
                base_url="https://open.bigmodel.cn/api/paas/v4"
            )

            prompt = f"""请仔细识别这张图片中的文字内容，并提取以下医学参数的值：
{', '.join(PARAMETERS)}

请以JSON格式返回结果，格式如下：
{{
    "IVS;d": "值",
    "LVID;d": "值",
    "LVPW;d": "值",
    "IVS;s": "值",
    "LVID;s": "值",
    "LVPW;s": "值",
    "LV Vol;d": "值",
    "LV Vol;s": "值",
    "EF": "值",
    "FS": "值",
    "LV Mass": "值",
    "LV Mass Cor": "值"
}}

如果某个参数在图片中没有找到，请将其值设为"N/A"。只返回JSON格式，不要有其他文本。"""

            response = client.chat.completions.create(
                model="glm-4v-flash",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "image_url", "image_url": {"url": base64_image}},
                            {"type": "text", "text": prompt}
                        ]
                    }
                ],
                thinking={"type": "enabled"}
            )

            result_text = response.choices[0].message.content

            try:
                json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', result_text, re.DOTALL)
                if json_match:
                    json_str = json_match.group(0)
                    return json.loads(json_str)
                else:
                    return json.loads(result_text)
            except json.JSONDecodeError:
                self.log_message(f"JSON解析失败\n", "warning")
                return None

        except Exception as e:
            self.log_message(f"调用智谱API失败: {e}\n", "error")
            return None

    def parse_api_response(self, response_data):
        """解析API返回的结果"""
        result = {param: "N/A" for param in PARAMETERS}

        if response_data is None:
            return result

        try:
            if isinstance(response_data, dict):
                for param in PARAMETERS:
                    if param in response_data:
                        result[param] = response_data[param]
            elif isinstance(response_data, str):
                for param in PARAMETERS:
                    pattern = rf"{re.escape(param)}[：:\s]*([0-9.,]+(?:\s*[a-zA-Z%]*)?)"
                    match = re.search(pattern, response_data, re.IGNORECASE)
                    if match:
                        result[param] = match.group(1).strip()
        except Exception as e:
            self.log_message(f"解析API响应失败: {e}\n", "error")

        return result

    def process_images(self):
        """处理图片"""
        # 验证 API Key
        if not self.api_key.get():
            messagebox.showwarning("警告", "请输入智谱 AI API Key")
            return

        # 获取文件夹列表
        folder_paths = self.folder_paths_list.copy()

        if not folder_paths:
            messagebox.showwarning("警告", "请至少选择一个文件夹")
            return

        # 如果正在处理，不重复执行
        if self.processing:
            return

        # 清空输出区域
        self.text_output.delete("1.0", tk.END)
        self.log_message("开始处理...\n\n", "info")

        # 设置处理状态
        self.set_processing_state(True)

        # 启动后台线程
        self.process_thread = threading.Thread(
            target=self.process_images_worker,
            args=(folder_paths,),
            daemon=True
        )
        self.process_thread.start()

    def process_images_worker(self, folder_paths):
        """后台线程处理图片"""
        try:
            start_time = time.time()
            self.log_message("start >>>>>\n\n", "info")

            # 支持的图片扩展名
            image_extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tif', '.tiff')

            # 收集所有图片文件
            all_image_files = []
            for folder_path in folder_paths:
                self.log_message(f"扫描文件夹: {folder_path}\n", "info")
                for root, dirs, files in os.walk(folder_path):
                    for file in files:
                        if file.lower().endswith(image_extensions):
                            all_image_files.append(os.path.join(root, file))

            if not all_image_files:
                self.log_message("未找到图片文件\n", "warning")
                self.root.after(0, lambda: self.set_processing_state(False))
                return

            self.log_message(f"找到 {len(all_image_files)} 个图片文件\n\n", "success")

            # 结果列表
            results = []

            # 处理每个图片文件
            for idx, image_path in enumerate(all_image_files, 1):
                self.log_message(f"[{idx}/{len(all_image_files)}] {os.path.basename(image_path)}\n", "info")

                # 编码图片
                base64_image, mime_type = self.encode_image_to_base64(image_path)
                if base64_image is None:
                    self.log_message(f"  跳过: 编码失败\n", "warning")
                    continue

                # 调用API
                self.log_message(f"  调用 API...\n", "info")
                response = self.call_glm4v_api(base64_image, mime_type)

                # 解析结果
                params = self.parse_api_response(response)

                # 添加文件名
                row_data = {
                    "文件名": os.path.basename(image_path),
                    "文件路径": image_path
                }
                row_data.update(params)

                results.append(row_data)
                self.log_message(f"  完成\n", "success")

            # 保存到Excel
            if results:
                output_path = self.output_file.get()
                # 确保输出目录存在
                output_dir = os.path.dirname(output_path)
                if output_dir:
                    os.makedirs(output_dir, exist_ok=True)

                df = pd.DataFrame(results)
                columns = ["文件名", "文件路径"] + PARAMETERS
                df = df[columns]

                df.to_excel(output_path, index=False, engine='openpyxl')

                # 使用 openpyxl 设置样式
                from openpyxl import load_workbook
                wb = load_workbook(output_path)
                ws = wb.active

                # 设置表头加粗
                bold_font = Font(bold=True)
                for cell in ws[1]:
                    cell.font = bold_font

                wb.save(output_path)

                end_time = time.time()
                elapsed_time = "{:.1f}".format(end_time - start_time)
                self.log_message(f"\n完成! 耗时 {elapsed_time}s\n", "success")
                self.log_message(f"结果已保存到: {output_path}\n", "success")
                self.log_message(f"共处理 {len(results)} 个图片文件\n", "info")
            else:
                self.log_message("\n没有成功处理任何图片文件\n", "warning")

            # 处理完成
            self.root.after(0, lambda: self.set_processing_state(False))
            self.root.after(0, lambda: messagebox.showinfo("完成", "OCR 识别完成！"))

        except Exception as e:
            self.log_message(f"\n处理出错: {str(e)}\n", "error")
            self.root.after(0, lambda: self.set_processing_state(False))
            self.root.after(0, lambda: messagebox.showerror("错误", f"处理失败: {str(e)}"))


if __name__ == "__main__":
    root = tk.Tk()
    app = OCRProcessorUI(root)
    root.mainloop()
