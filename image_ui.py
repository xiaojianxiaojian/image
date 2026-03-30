import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import imagej
import ctypes
from ctypes import wintypes
import sys
import cv2
import numpy as np
import shutil
import threading
from queue import Queue


# Windows API 多选文件夹对话框
class MultiFolderDialog:
    def __init__(self):
        if sys.platform != 'win32':
            raise RuntimeError("多选文件夹对话框仅支持 Windows")

        # 定义 COM 接口
        self.IID_IFileOpenDialog = ctypes.c_void_p.in_dll(
            ctypes.oledll.ole32, 'IID_IFileOpenDialog'
        )

        # 定义方法
        self.CoCreateInstance = ctypes.windll.ole32.CoCreateInstance
        self.CoInitialize = ctypes.windll.ole32.CoInitialize
        self.CoUninitialize = ctypes.windll.ole32.CoUninitialize
        self.CoTaskMemFree = ctypes.windll.ole32.CoTaskMemFree

    def select_folders(self, title="选择文件夹"):
        self.CoInitialize(None)

        try:
            # 使用标准方法，只能单选，所以使用递归方式
            root = tk.Tk()
            root.withdraw()
            folder = filedialog.askdirectory(title=title)
            root.destroy()

            if folder:
                # 检查是否包含子文件夹
                subfolders = []
                for item in os.listdir(folder):
                    item_path = os.path.join(folder, item)
                    if os.path.isdir(item_path):
                        subfolders.append(item_path)

                if subfolders:
                    # 询问是否添加所有子文件夹
                    msg = f"发现 {len(subfolders)} 个子文件夹，是否全部添加？\n\n"
                    msg += "是 - 添加所有子文件夹\n"
                    msg += "否 - 只添加选中的文件夹\n"
                    msg += "取消 - 不添加"
                    choice = messagebox.askyesnocancel("确认", msg)

                    if choice is None:
                        return []
                    elif choice:
                        return subfolders
                    else:
                        return [folder]
                else:
                    return [folder]
            return []

        finally:
            self.CoUninitialize()


def select_folders_multi(title="选择文件夹"):
    """使用递归方式模拟多选"""
    root = tk.Tk()
    root.withdraw()
    folders = []

    while True:
        folder = filedialog.askdirectory(title=title + " (可多次选择)")
        if not folder:
            break

        # 检查是否包含子文件夹
        subfolders = []
        for item in os.listdir(folder):
            item_path = os.path.join(folder, item)
            if os.path.isdir(item_path):
                subfolders.append(item_path)

        if subfolders:
            # 询问是否添加所有子文件夹
            msg = f"发现 {len(subfolders)} 个子文件夹，是否全部添加？\n\n"
            msg += "是 - 添加所有子文件夹\n"
            msg += "否 - 只添加选中的文件夹\n"
            msg += "取消 - 停止选择"
            choice = messagebox.askyesnocancel("确认", msg)

            if choice is None:
                break
            elif choice:
                folders.extend(subfolders)
            else:
                if folder not in folders:
                    folders.append(folder)
        else:
            if folder not in folders:
                folders.append(folder)

        # 询问是否继续选择
        if not messagebox.askyesno("继续", "是否继续选择其他文件夹？"):
            break

    root.destroy()
    return folders


# ============ 去除背景相关方法 ============

def remove_background(image_path, output_path, lower_color, upper_color):
    """去除图片背景并替换为白色"""
    # 读取图像
    image = cv2.imread(image_path)
    if image is None:
        raise ValueError(f"无法读取图片: {image_path}")

    # 转换为HSV颜色空间
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # 定义颜色范围以创建掩码
    lower_bound = np.array(lower_color)
    upper_bound = np.array(upper_color)

    # 创建掩码
    mask = cv2.inRange(hsv, lower_bound, upper_bound)

    # 反转掩码
    mask = cv2.bitwise_not(mask)

    # 应用掩码
    result = cv2.bitwise_and(image, image, mask=mask)

    # 创建白色背景图像
    white_background = np.ones_like(image, dtype=np.uint8) * 255

    # 将前景图像覆盖在白色背景上
    final_result = cv2.add(result, cv2.bitwise_and(white_background, white_background, mask=cv2.bitwise_not(mask)))

    # 保存结果图像
    cv2.imwrite(output_path, final_result)


def process_folder_remove_background(input_folder, output_folder, lower_color, upper_color):
    """处理文件夹中所有图片，去除背景"""
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    processed_files = []
    for filename in os.listdir(input_folder):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.tif', '.bmp')):
            input_path = os.path.join(input_folder, filename)
            output_path = os.path.join(output_folder, filename)
            remove_background(input_path, output_path, lower_color, upper_color)
            processed_files.append(output_path)

    return processed_files


# ============ UI 类 ============

class ImageProcessorUI:
    # 配色方案
    COLORS = {
        'bg_main': '#F5F7FA',           # 主背景色
        'bg_panel': '#FFFFFF',          # 面板背景
        'primary': '#4A90D9',           # 主色调 - 蓝色
        'primary_hover': '#357ABD',     # 主色调悬停
        'success': '#5CB85C',           # 成功色 - 绿色
        'success_hover': '#4AA84A',     # 成功色悬停
        'danger': '#E74C3C',            # 危险色 - 红色
        'danger_hover': '#D0342C',      # 危险色悬停
        'warning': '#F39C12',           # 警告色 - 橙色
        'warning_hover': '#E8910C',     # 警告色悬停
        'info': '#5BC0DE',              # 信息色 - 青色
        'text_primary': '#2C3E50',      # 主要文字
        'text_secondary': '#7F8C8D',    # 次要文字
        'border': '#E1E8ED',            # 边框色
        'border_light': '#ECF0F1',      # 浅边框
        'input_bg': '#FAFBFC',          # 输入框背景
        'shadow': 'rgba(0,0,0,0.1)',    # 阴影
    }

    def __init__(self, root):
        self.root = root
        self.root.title("图片处理工具")
        self.root.geometry("900x900")
        self.root.minsize(800, 600)
        self.root.configure(bg=self.COLORS['bg_main'])

        # 输出文件夹路径
        self.output_folder = tk.StringVar(value="D:/code/image/result/")

        # 去除背景配置
        self.enable_remove_bg = tk.BooleanVar(value=True)
        self.lower_color_h = tk.IntVar(value=0)
        self.lower_color_s = tk.IntVar(value=0)
        self.lower_color_v = tk.IntVar(value=200)
        self.upper_color_h = tk.IntVar(value=180)
        self.upper_color_s = tk.IntVar(value=180)
        self.upper_color_v = tk.IntVar(value=255)

        # 文件夹路径列表
        self.folder_paths_list = []

        # 线程相关
        self.processing = False
        self.process_thread = None
        self.log_queue = Queue()

        # 初始化 ImageJ
        self.ij = imagej.init(mode='headless')

        self.create_widgets()
        self.check_log_queue()

    def create_widgets(self):
        # 主容器 - 使用 grid 布局
        main_container = tk.Frame(self.root, bg=self.COLORS['bg_main'])
        main_container.pack(fill="both", expand=True, padx=16, pady=16)

        # 标题区域 - 带装饰效果
        title_frame = tk.Frame(main_container, bg=self.COLORS['bg_main'])
        title_frame.grid(row=0, column=0, columnspan=2, pady=(0, 12), sticky="ew")

        # 标题图标和文字
        title_container = tk.Frame(title_frame, bg=self.COLORS['primary'], height=50)
        title_container.pack(fill="x")
        title_container.pack_propagate(False)

        title_label = tk.Label(
            title_container,
            text="  🖼️  图片处理工具",
            font=("Microsoft YaHei", 16, "bold"),
            bg=self.COLORS['primary'],
            fg="white",
            anchor="w"
        )
        title_label.pack(fill="both", expand=True, padx=16, pady=8)

        # 左侧面板 - 配置区域
        left_panel = tk.Frame(main_container, bg=self.COLORS['bg_panel'], relief="flat", bd=0)
        left_panel.grid(row=1, column=0, sticky="nsew", padx=(0, 10))

        # 右侧面板 - 文件夹列表和日志
        right_panel = tk.Frame(main_container, bg=self.COLORS['bg_main'])
        right_panel.grid(row=1, column=1, sticky="nsew")

        # 配置 grid 权重
        main_container.columnconfigure(1, weight=1)
        main_container.rowconfigure(1, weight=1)

        # 左侧面板内容容器 - 使用 grid
        left_content = tk.Frame(left_panel, bg=self.COLORS['bg_panel'])
        left_content.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        left_panel.columnconfigure(0, weight=1)
        left_panel.rowconfigure(0, weight=1)

        # ============ 左侧面板 ============
        row_idx = 0

        # 去除背景配置区域 - 美化样式
        bg_frame = tk.LabelFrame(
            left_content,
            text=" 🎨 去除背景配置 (HSV) ",
            font=("Microsoft YaHei", 10, "bold"),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_primary'],
            relief="solid",
            bd=1,
            highlightbackground=self.COLORS['border'],
            highlightthickness=1
        )
        bg_frame.grid(row=row_idx, column=0, sticky="ew", pady=(0, 12))
        row_idx += 1

        # 启用去除背景复选框 - 自定义样式
        enable_cb = tk.Checkbutton(
            bg_frame,
            text="启用去除背景",
            variable=self.enable_remove_bg,
            font=("Microsoft YaHei", 10),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_primary'],
            selectcolor=self.COLORS['bg_panel'],
            activebackground=self.COLORS['bg_panel'],
            activeforeground=self.COLORS['primary'],
            cursor="hand2"
        )
        enable_cb.pack(anchor="w", padx=12, pady=(10, 8))

        # HSV 颜色范围配置 - 使用 grid 布局更紧凑
        hsv_inner = tk.Frame(bg_frame, bg=self.COLORS['bg_panel'])
        hsv_inner.pack(fill="x", padx=12, pady=(0, 12))

        # 表头 - 带背景色
        header_bg = self.COLORS['primary']
        tk.Label(hsv_inner, text="", width=8, bg=self.COLORS['bg_panel']).grid(row=0, column=0)
        tk.Label(hsv_inner, text="  H  ", font=("Microsoft YaHei", 9, "bold"), bg=header_bg, fg="white").grid(row=0, column=1, padx=1)
        tk.Label(hsv_inner, text="  S  ", font=("Microsoft YaHei", 9, "bold"), bg=header_bg, fg="white").grid(row=0, column=2, padx=1)
        tk.Label(hsv_inner, text="  V  ", font=("Microsoft YaHei", 9, "bold"), bg=header_bg, fg="white").grid(row=0, column=3, padx=1)

        # 输入框样式
        entry_style = {
            'justify': "center",
            'font': ("Microsoft YaHei", 10),
            'bg': self.COLORS['input_bg'],
            'fg': self.COLORS['text_primary'],
            'insertbackground': self.COLORS['primary'],
            'relief': "solid",
            'bd': 1,
            'highlightthickness': 1,
            'highlightbackground': self.COLORS['border'],
            'highlightcolor': self.COLORS['primary']
        }

        # Lower Color
        tk.Label(hsv_inner, text="Lower:", font=("Microsoft YaHei", 10), bg=self.COLORS['bg_panel'], fg=self.COLORS['text_secondary']).grid(row=1, column=0, sticky="w", pady=(8, 4))
        tk.Entry(hsv_inner, textvariable=self.lower_color_h, width=5, **entry_style).grid(row=1, column=1, padx=2, pady=(8, 4))
        tk.Entry(hsv_inner, textvariable=self.lower_color_s, width=5, **entry_style).grid(row=1, column=2, padx=2, pady=(8, 4))
        tk.Entry(hsv_inner, textvariable=self.lower_color_v, width=5, **entry_style).grid(row=1, column=3, padx=2, pady=(8, 4))

        # Upper Color
        tk.Label(hsv_inner, text="Upper:", font=("Microsoft YaHei", 10), bg=self.COLORS['bg_panel'], fg=self.COLORS['text_secondary']).grid(row=2, column=0, sticky="w", pady=(4, 8))
        tk.Entry(hsv_inner, textvariable=self.upper_color_h, width=5, **entry_style).grid(row=2, column=1, padx=2, pady=(4, 8))
        tk.Entry(hsv_inner, textvariable=self.upper_color_s, width=5, **entry_style).grid(row=2, column=2, padx=2, pady=(4, 8))
        tk.Entry(hsv_inner, textvariable=self.upper_color_v, width=5, **entry_style).grid(row=2, column=3, padx=2, pady=(4, 8))

        # 输出文件夹配置 - 美化样式
        output_frame = tk.LabelFrame(
            left_content,
            text=" 📁 输出设置 ",
            font=("Microsoft YaHei", 10, "bold"),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_primary'],
            relief="solid",
            bd=1,
            highlightbackground=self.COLORS['border'],
            highlightthickness=1
        )
        output_frame.grid(row=row_idx, column=0, sticky="ew", pady=(0, 12))
        row_idx += 1

        output_label = tk.Label(output_frame, text="输出文件夹:", font=("Microsoft YaHei", 10), bg=self.COLORS['bg_panel'], fg=self.COLORS['text_secondary'])
        output_label.pack(anchor="w", padx=12, pady=(10, 6))

        output_entry_frame = tk.Frame(output_frame, bg=self.COLORS['bg_panel'])
        output_entry_frame.pack(fill="x", padx=12, pady=(0, 10))

        output_entry = tk.Entry(
            output_entry_frame,
            textvariable=self.output_folder,
            font=("Microsoft YaHei", 10),
            bg=self.COLORS['input_bg'],
            fg=self.COLORS['text_primary'],
            insertbackground=self.COLORS['primary'],
            relief="solid",
            bd=1,
            highlightthickness=1,
            highlightbackground=self.COLORS['border'],
            highlightcolor=self.COLORS['primary']
        )
        output_entry.pack(side="left", fill="x", expand=True)

        btn_browse = tk.Button(
            output_entry_frame,
            text="浏览",
            command=self.browse_output_folder,
            width=6,
            font=("Microsoft YaHei", 9),
            bg=self.COLORS['info'],
            fg="white",
            cursor="hand2",
            relief="flat",
            activebackground=self.COLORS['primary'],
            activeforeground="white"
        )
        btn_browse.pack(side="left", padx=(6, 0))

        # 操作按钮区域 - 美化样式
        action_frame = tk.Frame(left_content, bg=self.COLORS['bg_panel'])
        action_frame.grid(row=row_idx, column=0, sticky="ew", pady=(0, 12))
        row_idx += 1

        # 创建渐变效果的按钮
        self.btn_process = tk.Button(
            action_frame,
            text="▶ 开始处理",
            command=self.process_images,
            bg=self.COLORS['success'],
            fg="white",
            font=("Microsoft YaHei", 12, "bold"),
            height=2,
            cursor="hand2",
            relief="flat",
            activebackground=self.COLORS['success_hover'],
            activeforeground="white",
            bd=0
        )
        self.btn_process.pack(fill="x", pady=(4, 4))

        # 添加说明文字
        info_label = tk.Label(
            left_content,
            text="💡 提示: 可以添加多个文件夹进行批量处理",
            font=("Microsoft YaHei", 9),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_secondary']
        )
        info_label.grid(row=row_idx, column=0, sticky="w", pady=(0, 8))

        # ============ 右侧面板 ============
        right_row = 0

        # 文件夹选择区域 - 美化样式
        folder_frame = tk.LabelFrame(
            right_panel,
            text=" 📂 输入文件夹 (支持多选) ",
            font=("Microsoft YaHei", 10, "bold"),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_primary'],
            relief="solid",
            bd=1,
            highlightbackground=self.COLORS['border'],
            highlightthickness=1
        )
        folder_frame.grid(row=right_row, column=0, sticky="nsew", pady=(0, 10))
        right_row += 1

        right_panel.rowconfigure(right_row - 1, weight=1)
        right_panel.columnconfigure(0, weight=1)

        # 列表框容器 - 美化
        listbox_container = tk.Frame(folder_frame, bg=self.COLORS['bg_panel'])
        listbox_container.pack(fill="both", expand=True, padx=10, pady=(8, 8))

        scrollbar = tk.Scrollbar(
            listbox_container,
            bg=self.COLORS['bg_panel'],
            troughcolor=self.COLORS['border_light'],
            activebackground=self.COLORS['primary']
        )
        scrollbar.pack(side="right", fill="y")

        self.folder_listbox = tk.Listbox(
            listbox_container,
            font=("Microsoft YaHei", 10),
            yscrollcommand=scrollbar.set,
            selectmode=tk.MULTIPLE,
            bg=self.COLORS['input_bg'],
            fg=self.COLORS['text_primary'],
            selectbackground=self.COLORS['primary'],
            selectforeground="white",
            relief="solid",
            bd=1,
            highlightthickness=1,
            highlightbackground=self.COLORS['border'],
            highlightcolor=self.COLORS['primary']
        )
        self.folder_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.folder_listbox.yview)

        # 按钮区域 - 美化
        btn_frame = tk.Frame(folder_frame, bg=self.COLORS['bg_panel'])
        btn_frame.pack(fill="x", padx=10, pady=(0, 10))

        btn_style_common = {
            "font": ("Microsoft YaHei", 10),
            "cursor": "hand2",
            "relief": "flat",
            "bd": 0,
            "height": 1
        }

        btn_add = tk.Button(
            btn_frame,
            text="➕ 添加文件夹",
            command=self.add_folder,
            bg=self.COLORS['primary'],
            fg="white",
            activebackground=self.COLORS['primary_hover'],
            activeforeground="white",
            **btn_style_common
        )
        btn_add.pack(side="left", fill="x", expand=True, padx=(0, 4))

        btn_remove = tk.Button(
            btn_frame,
            text="➖ 移除选中",
            command=self.remove_folder,
            bg=self.COLORS['danger'],
            fg="white",
            activebackground=self.COLORS['danger_hover'],
            activeforeground="white",
            **btn_style_common
        )
        btn_remove.pack(side="left", fill="x", expand=True, padx=4)

        btn_clear = tk.Button(
            btn_frame,
            text="🗑️ 清空全部",
            command=self.clear_folders,
            bg=self.COLORS['warning'],
            fg="white",
            activebackground=self.COLORS['warning_hover'],
            activeforeground="white",
            **btn_style_common
        )
        btn_clear.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # 输出日志区域 - 美化样式
        log_frame = tk.LabelFrame(
            right_panel,
            text=" 📋 输出日志 ",
            font=("Microsoft YaHei", 10, "bold"),
            bg=self.COLORS['bg_panel'],
            fg=self.COLORS['text_primary'],
            relief="solid",
            bd=1,
            highlightbackground=self.COLORS['border'],
            highlightthickness=1
        )
        log_frame.grid(row=right_row, column=0, sticky="nsew")
        right_row += 1

        right_panel.rowconfigure(right_row - 1, weight=1)

        # 配置日志文本标签颜色
        self.text_output = scrolledtext.ScrolledText(
            log_frame,
            font=("Consolas", 10),
            wrap="word",
            bg="#1E1E1E",
            fg="#D4D4D4",
            insertbackground="#FFFFFF",
            relief="solid",
            bd=1,
            highlightthickness=0
        )

        # 配置日志颜色标签
        self.text_output.tag_config("info", foreground="#5BC0DE")
        self.text_output.tag_config("success", foreground="#5CB85C")
        self.text_output.tag_config("error", foreground="#E74C3C")
        self.text_output.tag_config("warning", foreground="#F39C12")

        self.text_output.pack(fill="both", expand=True, padx=10, pady=(8, 10))

    # ============ 线程相关方法 ============

    def log_message(self, message, tag=None):
        """线程安全的日志输出方法"""
        self.log_queue.put((message, tag))

    def check_log_queue(self):
        """检查日志队列并更新UI（在主线程中调用）"""
        try:
            while True:
                message, tag = self.log_queue.get_nowait()
                if tag:
                    self.text_output.insert(tk.END, message, tag)
                else:
                    self.text_output.insert(tk.END, message)
                self.text_output.see(tk.END)
        except:
            pass
        # 每100ms检查一次
        self.root.after(100, self.check_log_queue)

    def set_processing_state(self, processing):
        """设置处理状态，启用/禁用控件"""
        self.processing = processing
        if processing:
            self.btn_process.config(text="⏳ 处理中...", state="disabled", bg=self.COLORS['text_secondary'])
        else:
            self.btn_process.config(text="▶ 开始处理", state="normal", bg=self.COLORS['success'])

    def add_folder(self):
        folders = select_folders_multi(title="选择文件夹")
        for folder in folders:
            if folder not in self.folder_paths_list:
                self.folder_paths_list.append(folder)
                self.folder_listbox.insert(tk.END, folder)

    def remove_folder(self):
        selected = self.folder_listbox.curselection()
        if selected:
            # 倒序删除，避免索引变化
            for index in reversed(selected):
                self.folder_listbox.delete(index)
                del self.folder_paths_list[index]

    def clear_folders(self):
        self.folder_listbox.delete(0, tk.END)
        self.folder_paths_list.clear()

    def browse_output_folder(self):
        folder = filedialog.askdirectory(title="选择输出文件夹")
        if folder:
            # 确保路径以/结尾
            if not folder.endswith('/') and not folder.endswith('\\'):
                folder = folder + '/'
            self.output_folder.set(folder)

    def remove_background_folder(self, folder_path, temp_output_folder):
        """去除文件夹中所有图片的背景"""
        lower_color = [
            self.lower_color_h.get(),
            self.lower_color_s.get(),
            self.lower_color_v.get()
        ]
        upper_color = [
            self.upper_color_h.get(),
            self.upper_color_s.get(),
            self.upper_color_v.get()
        ]

        # 创建临时输出文件夹
        if not os.path.exists(temp_output_folder):
            os.makedirs(temp_output_folder)

        # 处理文件夹中的所有图片
        for filename in os.listdir(folder_path):
            if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.tif', '.bmp')):
                input_path = os.path.join(folder_path, filename)
                output_path = os.path.join(temp_output_folder, filename)
                try:
                    remove_background(input_path, output_path, lower_color, upper_color)
                except Exception as e:
                    self.log_message(f"  错误: {filename} - {str(e)}\n", "error")

        return temp_output_folder

    def process_images_inner(self, folder_path):
        """处理单个文件夹的图片（调用原始50-106行逻辑）"""
        self.log_message(f"计算区域处理文件夹: {folder_path}\n", "info")

        results = []

        for file in os.listdir(folder_path):
            file_path_full = os.path.join(folder_path, file)
            if os.path.isfile(file_path_full):
                try:
                    image = self.ij.io().open(file_path_full)

                    # 将 ImageJ 图像转换为 NumPy 数组
                    image_array = self.ij.py.from_java(image)

                    # 提取蓝色通道
                    blue_channel = image_array[:, :, 2]

                    # 创建仅包含蓝色通道的图像
                    processed_image = self.ij.py.to_java(blue_channel)

                    output_path = 'output.tif'
                    self.ij.io().save(processed_image, output_path)

                    image_out = self.ij.io().open(output_path)

                    imp_default = self.ij.py.to_imageplus(image_out)
                    imp_all = self.ij.py.to_imageplus(image_out)

                    self.ij.IJ.run("Set Measurements...", "area_fraction")
                    self.ij.IJ.setAutoThreshold(imp_default, "Default")
                    self.ij.IJ.setRawThreshold(imp_all, 0, 254)

                    output_default = self.ij.IJ.getValue(imp_default, "%Area")
                    output_all = self.ij.IJ.getValue(imp_all, "%Area")

                    formatted_number = "{:.2f}".format((float(output_all) - float(output_default)) / float(output_all))

                    results.append({
                        "Folder": folder_path,
                        "File": file,
                        "Value": formatted_number
                    })

                    os.remove(output_path)

                except Exception as e:
                    self.log_message(f"错误: {file} - {str(e)}\n", "error")

        return results

    def process_images(self):
        # 从列表框获取文件夹路径
        folder_paths = self.folder_paths_list.copy()

        if not folder_paths:
            messagebox.showwarning("警告", "请至少选择一个文件夹")
            return

        # 如果正在处理，不重复执行
        if self.processing:
            return

        # 清空输出区域
        self.text_output.delete("1.0", tk.END)
        self.log_message("开始处理...\n\n", "info")

        # 设置处理状态
        self.set_processing_state(True)

        # 启动后台线程
        self.process_thread = threading.Thread(
            target=self.process_images_worker,
            args=(folder_paths,),
            daemon=True
        )
        self.process_thread.start()

    def process_images_worker(self, folder_paths):
        """后台线程处理图片"""
        try:
            start_time = time.time()
            self.log_message("start >>>>>\n\n", "info")

            # 收集结果，按文件夹分组
            folder_results = {}

            # 创建临时文件夹用于存储去除背景后的图片
            temp_base = os.path.join(os.path.dirname(self.output_folder.get()), "temp_processed")
            if os.path.exists(temp_base):
                shutil.rmtree(temp_base)
            os.makedirs(temp_base)

            try:
                for folder_path in folder_paths:
                    # 获取文件夹名称
                    folder_name = os.path.basename(folder_path.rstrip('/\\'))

                    # 如果启用去除背景，先处理背景
                    if self.enable_remove_bg.get():
                        temp_output_folder = os.path.join(temp_base, folder_name)
                        processed_folder = self.remove_background_folder(folder_path, temp_output_folder)
                        self.log_message(f"去除背景处理文件夹: {processed_folder}\n", "info")
                        results = self.process_images_inner(processed_folder)
                    else:
                        results = self.process_images_inner(folder_path)

                    folder_results[folder_path] = results

                    end_time = time.time()
                    elapsed_time = "{:.1f}".format(end_time - start_time)
                    self.log_message(f"\nstop <<<<<<<<<<<< costs {elapsed_time}s\n\n", "info")

                # 如果启用去除背景，保存处理后的图片到输出文件夹
                if self.enable_remove_bg.get():
                    self.log_message("\n保存去除背景后的图片...\n", "info")
                    for folder_path in folder_paths:
                        folder_name = os.path.basename(folder_path.rstrip('/\\'))
                        temp_folder = os.path.join(temp_base, folder_name)
                        if os.path.exists(temp_folder):
                            output_folder = self.output_folder.get() + folder_name + "_no_bg/"
                            if os.path.exists(output_folder):
                                shutil.rmtree(output_folder)
                            shutil.copytree(temp_folder, output_folder)
                            self.log_message(f"已保存到: {output_folder}\n", "success")

            finally:
                # 清理临时文件夹
                if os.path.exists(temp_base):
                    shutil.rmtree(temp_base)

            # 写入Excel
            if folder_results:
                output_excel = self.output_folder.get() + "result.xlsx"
                wb = Workbook()
                ws = wb.active
                ws.title = "Result"

                # 创建粗体字体
                bold_font = Font(bold=True)

                headers = ["Folder", "File", "Value"]
                current_row = 1

                for folder_path, results in folder_results.items():
                    # 写入列名（表头加粗）
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=header)
                        cell.font = bold_font
                    current_row += 1

                    # 计算平均值
                    values = [float(r["Value"]) for r in results]
                    avg_value = sum(values) / len(values) if values else 0
                    folder_name = os.path.basename(folder_path)

                    # 写入数据
                    for result in results:
                        ws.cell(row=current_row, column=1, value=folder_name)
                        ws.cell(row=current_row, column=2, value=result["File"])
                        ws.cell(row=current_row, column=3, value=float(result["Value"]))
                        current_row += 1

                    # 写入平均值行（平均值加粗）
                    avg_cell1 = ws.cell(row=current_row, column=1, value="平均值")
                    avg_cell1.font = bold_font
                    avg_cell3 = ws.cell(row=current_row, column=3, value=round(avg_value, 2))
                    avg_cell3.font = bold_font
                    current_row += 1

                    # 空白行
                    current_row += 1

                wb.save(output_excel)
                self.log_message(f"\nExcel已保存到: {output_excel}\n", "success")

            # 处理完成，恢复UI状态
            self.root.after(0, lambda: self.set_processing_state(False))
            self.root.after(0, lambda: messagebox.showinfo("完成", "处理完成！"))

        except Exception as e:
            self.log_message(f"\n处理出错: {str(e)}\n", "error")
            self.root.after(0, lambda: self.set_processing_state(False))
            self.root.after(0, lambda: messagebox.showerror("错误", f"处理失败: {str(e)}"))


if __name__ == "__main__":
    root = tk.Tk()
    app = ImageProcessorUI(root)
    root.mainloop()
